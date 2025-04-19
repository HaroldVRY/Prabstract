import os
import time
import pdfplumber
import openai
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.worksheet.table import Table, TableStyleInfo
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.mime.text import MIMEText

# Configura tu API Key de OpenAI
openai.api_key = 'COPIA TU API KEY AQUÍ'

# Ruta a ChromeDriver en tu sistema
driver_path = r'd:\Users\User\Desktop\chromedriver-win64\chromedriver.exe'
download_dir = os.path.join(os.getcwd(), 'downloads')

# Crear una carpeta de descargas si no existe
if not os.path.exists(download_dir):
    os.makedirs(download_dir)

# Configuración de las opciones del navegador
prefs = {
    "download.default_directory": download_dir,
    "download.prompt_for_download": False,
    "plugins.always_open_pdf_externally": True
}
options = webdriver.ChromeOptions()
options.add_experimental_option("prefs", prefs)

# Configurar el servicio de ChromeDriver
service = Service(executable_path=driver_path)
driver = webdriver.Chrome(service=service, options=options)

# Función para cerrar el pop-up o hacer clic fuera de él
def cerrar_popup():
    try:
        ActionChains(driver).move_by_offset(10, 10).click().perform()
        print("Pop-up cerrado exitosamente.")
    except Exception as e:
        print(f"No se pudo cerrar el pop-up: {str(e)}")

# Función para descargar solo los PDFs visibles con la clase 'destacados'
def descargar_pdfs_clase_destacados():
    url = "https://www.pj.gob.pe/wps/wcm/connect/CorteSuprema/s_cortes_suprema_home/as_Inicio/"
    driver.get(url)
    cerrar_popup()
    
    # Obtener solo los enlaces que tienen la clase 'destacados'
    enlaces_pdf = driver.find_elements(By.CLASS_NAME, "destacados")
    urls = [enlace.get_attribute('href') for enlace in enlaces_pdf if enlace.get_attribute('href')]

    # Descargar cada enlace encontrado usando sus URLs
    for url in urls:
        if url:
            try:
                driver.execute_script("window.open('');")
                driver.switch_to.window(driver.window_handles[1])
                driver.get(url)
                print(f"Descargando desde: {url}")
                time.sleep(5)
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
            except Exception as e:
                print(f"No se pudo descargar el archivo desde {url}. Error: {str(e)}")

# Función para extraer texto del PDF usando pdfplumber
def extraer_texto_pdf(ruta_pdf):
    with pdfplumber.open(ruta_pdf) as pdf:
        texto_completo = ""
        for pagina in pdf.pages:
            texto_completo += pagina.extract_text()
        print(f"Texto extraído del PDF {ruta_pdf}:\n{texto_completo[:500]}...")
        return texto_completo

# Función para clasificar la categoría según palabras clave
def clasificar_categoria(texto):
    if "contrato" in texto.lower():
        return "Departamento de Contratos"
    elif "litigio" in texto.lower() or "apelación" in texto.lower() or "juicio" in texto.lower():
        return "Departamento de Litigios"
    elif "propiedad intelectual" in texto.lower() or "patente" in texto.lower() or "marca registrada" in texto.lower():
        return "Departamento de Propiedad Intelectual"
    elif "cumplimiento" in texto.lower() or "normativa" in texto.lower():
        return "Departamento de Cumplimiento"
    elif "laboral" in texto.lower() or "trabajo" in texto.lower():
        return "Departamento de Recursos Humanos"
    elif "corporativo" in texto.lower() or "accionistas" in texto.lower():
        return "Departamento Corporativo"
    elif "fiscal" in texto.lower() or "tributo" in texto.lower() or "impuesto" in texto.lower():
        return "Departamento Financiero"
    elif "ambiental" in texto.lower():
        return "Departamento de Responsabilidad Social"
    elif "penal" in texto.lower():
        return "Departamento de Litigios Penales"
    elif "regulación financiera" in texto.lower() or "mercado de valores" in texto.lower():
        return "Departamento Financiero"
    else:
        return "Departamento General"

# Función para llamar a la API de OpenAI para extraer datos
def extraer_datos_con_api(texto):
    prompt = f"Por favor, extrae la siguiente información del texto: \n\n{texto}\n\n" \
             "Dame el nombre, el ponente, la fecha (en formato DD/MM/YYYY), el lugar, la categoría y un resumen de al menos 150 caracteres."
    
    try:
        respuesta = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[{"role": "user", "content": prompt}]
        )
        
        resultado = respuesta.choices[0].message['content']
        print(f"Resultado de OpenAI:\n{resultado}")
        return resultado

    except Exception as e:
        print(f"Error al llamar a la API de OpenAI: {str(e)}")
        return None

# Función para procesar varios PDFs y extraer datos
def procesar_pdfs(archivo_excel):
    datos_extraidos = []

    for archivo_pdf in os.listdir(download_dir):
        if archivo_pdf.endswith(".pdf"):
            ruta_pdf = os.path.join(download_dir, archivo_pdf)
            texto = extraer_texto_pdf(ruta_pdf)
            resultado = extraer_datos_con_api(texto)

            if resultado:
                try:
                    datos = resultado.split('\n')
                    registro = {
                        "Nombre": datos[0].split(":")[1].strip(),
                        "Ponente": datos[1].split(":")[1].strip(),
                        "Fecha": pd.to_datetime(datos[2].split(":")[1].strip(), dayfirst=True).date(),
                        "Lugar": datos[3].split(":")[1].strip(),
                        "Categoría": clasificar_categoria(texto),
                        "Resumen": datos[5].split(":")[1].strip()
                    }
                    datos_extraidos.append(registro)
                except (IndexError, ValueError) as e:
                    print(f"Error al procesar los datos extraídos: {str(e)}")

    df = pd.DataFrame(datos_extraidos)

    try:
        with pd.ExcelWriter(archivo_excel, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name='Datos', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Datos']
            
            # Convertir a tabla de Excel
            table = Table(displayName="DatosTabla", ref=f"A1:F{len(df) + 1}")
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True
            )
            table.tableStyleInfo = style
            worksheet.add_table(table)
            
            # Ajustar el ancho de las columnas
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column_cells)
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width
            
            print("Los datos han sido procesados y guardados en el archivo Excel con formato de tabla.")
    except Exception as e:
        print(f"Error al escribir en el archivo Excel: {str(e)}")

# Función para crear gráficos y agregarlos al archivo Excel
def agregar_graficos_al_excel(archivo_excel):
    df = pd.read_excel(archivo_excel, sheet_name='Datos')
    
    # Crear el gráfico de distribución por categoría
    plt.figure(figsize=(8, 6))
    df['Categoría'].value_counts().plot(kind='bar')
    plt.title('Distribución por Categoría')
    plt.xlabel('Categoría')
    plt.ylabel('Cantidad')
    plt.tight_layout()
    plt.savefig("categoria_distribucion.png")
    plt.close()
    
    # Crear el gráfico de distribución por fecha
    plt.figure(figsize=(8, 6))
    df['Fecha'].value_counts().sort_index().plot(kind='line', marker='o')
    plt.title('Distribución de Documentos por Fecha')
    plt.xlabel('Fecha')
    plt.ylabel('Cantidad')
    plt.tight_layout()
    plt.savefig("fecha_distribucion.png")
    plt.close()

    # Agregar los gráficos al archivo Excel
    workbook = load_workbook(archivo_excel)
    
    # Agregar la hoja para el gráfico de Categoría
    sheet_categoria = workbook.create_sheet(title='Grafico_Categoria')
    img_categoria = Image("categoria_distribucion.png")
    sheet_categoria.add_image(img_categoria, "A1")
    
    # Agregar la hoja para el gráfico de Fecha
    sheet_fecha = workbook.create_sheet(title='Grafico_Fecha')
    img_fecha = Image("fecha_distribucion.png")
    sheet_fecha.add_image(img_fecha, "A1")
    
    workbook.save(archivo_excel)
    print("Gráficos agregados al archivo Excel.")

# Función para enviar el archivo por correo
def enviar_correo(contrasena, destinatario, archivo_adjunto, remitente="harold.reyna.y@uni.pe"):
    asunto = "Resultados de los documentos destacados categorizados"
    mensaje = """
    Se hace el envío de los nombres de los archivos encomendados ya categorizados para su posterior envío a las áreas respectivas de la empresa para su posterior análisis, adicionalmente se presenta la cantidad de documentos por fecha y por categoría.
    
    Gracias por su atención.
    Atentamente,
    Reyna Yangali Harold Victor
    Tel: 914-027-388
    """

    msg = MIMEMultipart()
    msg['From'] = remitente
    msg['To'] = destinatario
    msg['Subject'] = asunto
    msg.attach(MIMEText(mensaje, 'plain'))
    
    # Adjuntar el archivo
    with open(archivo_adjunto, "rb") as adjunto:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(adjunto.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename= {os.path.basename(archivo_adjunto)}")
        msg.attach(part)
    
    try:
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(remitente, contrasena)
            server.sendmail(remitente, destinatario, msg.as_string())
        print(f"Correo enviado exitosamente a {destinatario}.")
    except Exception as e:
        print(f"Error al enviar el correo: {e}")

# Función principal para ejecutar la RPA
def ejecutar_rpa():
    print(f"Los PDFs se descargarán en: {download_dir}")
    descargar_pdfs_clase_destacados()
    archivo_excel = 'resultados.xlsx'
    procesar_pdfs(archivo_excel)
    agregar_graficos_al_excel(archivo_excel)
    
    # Enviar el archivo por correo
    contrasena = "harolmazna123"  # Reemplaza con la contraseña de tu correo
    destinatario = "haroldreynayangali@gmail.com"
    enviar_correo(contrasena, destinatario, archivo_excel)
    
    driver.quit()  # Cerrar el navegador

# Ejecutar la RPA
ejecutar_rpa()
